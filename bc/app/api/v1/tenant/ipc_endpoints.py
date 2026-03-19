"""
OMMS IPC API Endpoints
نظام المستخلصات — جدول الكميات، إنشاء المستخلصات، تصدير PDF/Excel، الإشعارات
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import date, timedelta, datetime
import io, json, calendar

router = APIRouter(prefix="/ipc", tags=["📋 IPC — المستخلصات"])


def _get_db() -> Session:
    """Placeholder — replaced by tenant middleware in production"""
    from app.core.database import SessionLocal
    return SessionLocal()


# ════════════════════════════════════════════════════════
# IPC SETTINGS & FREQUENCY CONFIG
# ════════════════════════════════════════════════════════

@router.get("/projects/{project_id}/settings")
def get_ipc_settings(project_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import ProjectIPCSettings
    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    if not settings:
        raise HTTPException(404, "IPC settings not configured for this project")
    return settings


@router.post("/projects/{project_id}/settings")
def configure_ipc_settings(project_id: int, data: dict = Body(...), db: Session = Depends(_get_db)):
    """Configure IPC schedule and financial settings for a project"""
    from app.models.tenant.ipc_models import ProjectIPCSettings

    existing = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()

    if existing:
        for k, v in data.items():
            if hasattr(existing, k):
                setattr(existing, k, v)
        db.commit()
        db.refresh(existing)
        return existing

    settings = ProjectIPCSettings(
        project_id=project_id,
        **{k: v for k, v in data.items() if hasattr(ProjectIPCSettings, k)}
    )
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


@router.get("/projects/{project_id}/schedule")
def get_ipc_schedule(
    project_id: int,
    year: int = Query(default=None),
    db: Session = Depends(_get_db),
):
    """Get the full IPC schedule for a project"""
    from app.models.tenant.ipc_models import ProjectIPCSettings, IPC
    from app.models.tenant.models import Project

    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    project = db.query(Project).filter(Project.id == project_id).first()

    if not settings or not project:
        raise HTTPException(404, "Project or settings not found")

    year = year or date.today().year
    schedule = _generate_ipc_schedule(settings, project, year)

    # Mark which ones already exist
    existing_ipcs = db.query(IPC).filter(IPC.project_id == project_id).all()
    existing_periods = {(i.period_from.isoformat(), i.period_to.isoformat()): i for i in existing_ipcs}

    for slot in schedule:
        key = (slot["period_from"], slot["period_to"])
        ipc = existing_periods.get(key)
        slot["ipc_id"] = ipc.id if ipc else None
        slot["ipc_number"] = ipc.ipc_number if ipc else None
        slot["status"] = ipc.status if ipc else "not_created"
        slot["total_due"] = ipc.total_due if ipc else 0

    return {
        "project_id": project_id,
        "project_name": project.name_ar,
        "frequency": settings.frequency,
        "year": year,
        "schedule": schedule,
        "total_slots": len(schedule),
        "created": sum(1 for s in schedule if s["ipc_id"]),
        "pending": sum(1 for s in schedule if s["status"] in ["not_created", "draft"]),
    }


def _generate_ipc_schedule(settings, project, year: int) -> list:
    """Generate list of IPC periods based on frequency"""
    from app.models.tenant.ipc_models import IPCFrequency

    schedule = []
    proj_start = project.start_date or date(year, 1, 1)
    proj_end = project.end_date or date(year, 12, 31)
    first_date = settings.first_ipc_date or proj_start

    freq = settings.frequency
    cutoff = settings.cutoff_day or 25
    submit_day = settings.submission_day or 28

    if freq == IPCFrequency.monthly:
        # Monthly: from first_date month until project end
        current = first_date.replace(day=1)
        ipc_num = 1
        while current <= proj_end:
            if current.year == year:
                # Period: 1st to cutoff_day of month
                period_from = current
                _, last_day = calendar.monthrange(current.year, current.month)
                period_to = current.replace(day=min(cutoff, last_day))
                # Due date: submit_day of same month
                due_date = current.replace(day=min(submit_day, last_day))
                # Days remaining
                today = date.today()
                days_remaining = (due_date - today).days

                schedule.append({
                    "ipc_sequence": ipc_num,
                    "ipc_number": f"IPC-{ipc_num:03d}",
                    "period_from": period_from.isoformat(),
                    "period_to": period_to.isoformat(),
                    "cutoff_date": period_to.isoformat(),
                    "due_date": due_date.isoformat(),
                    "month_name": _month_name_ar(current.month),
                    "days_remaining": days_remaining,
                    "is_overdue": days_remaining < 0 and date.today() > due_date,
                    "is_due_soon": 0 <= days_remaining <= settings.notify_days_before,
                })
            next_month = current.month + 1
            next_year = current.year + (1 if next_month > 12 else 0)
            current = current.replace(year=next_year, month=(next_month - 1) % 12 + 1, day=1)
            ipc_num += 1

    elif freq == IPCFrequency.quarterly:
        quarters = [(1, 3), (4, 6), (7, 9), (10, 12)]
        for i, (start_m, end_m) in enumerate(quarters, 1):
            if proj_start <= date(year, start_m, 1) <= proj_end:
                _, last_day = calendar.monthrange(year, end_m)
                schedule.append({
                    "ipc_sequence": i,
                    "ipc_number": f"IPC-Q{i}-{year}",
                    "period_from": date(year, start_m, 1).isoformat(),
                    "period_to": date(year, end_m, last_day).isoformat(),
                    "cutoff_date": date(year, end_m, min(cutoff, last_day)).isoformat(),
                    "due_date": date(year, end_m, min(submit_day, last_day)).isoformat(),
                    "month_name": f"الربع {i} — {year}",
                    "days_remaining": (date(year, end_m, min(submit_day, last_day)) - date.today()).days,
                })

    elif freq == IPCFrequency.biweekly:
        current = first_date
        seq = 1
        while current <= proj_end and current.year == year:
            period_to = current + timedelta(days=13)
            due = period_to + timedelta(days=3)
            schedule.append({
                "ipc_sequence": seq,
                "ipc_number": f"IPC-{seq:03d}",
                "period_from": current.isoformat(),
                "period_to": period_to.isoformat(),
                "cutoff_date": period_to.isoformat(),
                "due_date": due.isoformat(),
                "month_name": f"فترة {seq} — {_month_name_ar(current.month)}",
                "days_remaining": (due - date.today()).days,
            })
            current = period_to + timedelta(days=1)
            seq += 1

    return schedule


def _month_name_ar(month: int) -> str:
    names = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
             "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
    return names[month - 1] if 1 <= month <= 12 else ""


# ════════════════════════════════════════════════════════
# BOQ MANAGEMENT
# ════════════════════════════════════════════════════════

@router.get("/projects/{project_id}/boq")
def get_boq(project_id: int, db: Session = Depends(_get_db)):
    """Get full BOQ with sections and items"""
    from app.models.tenant.ipc_models import ProjectIPCSettings, BOQSection, BOQItem

    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    if not settings:
        raise HTTPException(404, "IPC settings not found")

    sections = db.query(BOQSection).filter(
        BOQSection.settings_id == settings.id,
        BOQSection.is_active == True,
    ).order_by(BOQSection.sort_order).all()

    result = []
    total_contract = 0
    for section in sections:
        items = db.query(BOQItem).filter(
            BOQItem.section_id == section.id
        ).order_by(BOQItem.sort_order).all()

        section_total = sum(i.total_amount for i in items)
        total_contract += section_total

        result.append({
            "id": section.id,
            "code": section.code,
            "title": section.title,
            "title_ar": section.title_ar,
            "section_total": section_total,
            "items": [
                {
                    "id": item.id,
                    "item_number": item.item_number,
                    "description": item.description,
                    "description_ar": item.description_ar,
                    "unit": item.unit,
                    "quantity": item.quantity,
                    "unit_rate": item.unit_rate,
                    "total_amount": item.total_amount,
                    "previous_qty": item.previous_qty,
                    "completion_pct": round((item.previous_qty / item.quantity * 100) if item.quantity else 0, 1),
                }
                for item in items
            ],
        })

    return {
        "project_id": project_id,
        "sections": result,
        "total_sections": len(result),
        "total_items": sum(len(s["items"]) for s in result),
        "total_contract_value": total_contract,
    }


@router.post("/projects/{project_id}/boq/section")
def add_boq_section(project_id: int, data: dict = Body(...), db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import ProjectIPCSettings, BOQSection
    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    if not settings:
        raise HTTPException(404)
    count = db.query(BOQSection).filter(BOQSection.settings_id == settings.id).count()
    section = BOQSection(
        settings_id=settings.id,
        sort_order=count,
        **{k: v for k, v in data.items() if hasattr(BOQSection, k)}
    )
    db.add(section)
    db.commit()
    db.refresh(section)
    return section


@router.post("/boq/sections/{section_id}/items")
def add_boq_item(section_id: int, data: dict = Body(...), db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import BOQItem
    count = db.query(BOQItem).filter(BOQItem.section_id == section_id).count()
    qty = float(data.get("quantity", 0))
    rate = float(data.get("unit_rate", 0))
    item = BOQItem(
        section_id=section_id,
        sort_order=count,
        quantity=qty,
        unit_rate=rate,
        total_amount=round(qty * rate, 2),
        **{k: v for k, v in data.items()
           if hasattr(BOQItem, k) and k not in ("quantity", "unit_rate")}
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/boq/items/{item_id}")
def update_boq_item(item_id: int, data: dict = Body(...), db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import BOQItem
    item = db.query(BOQItem).filter(BOQItem.id == item_id).first()
    if not item:
        raise HTTPException(404)
    for k, v in data.items():
        if hasattr(item, k):
            setattr(item, k, v)
    item.total_amount = round(item.quantity * item.unit_rate, 2)
    db.commit()
    db.refresh(item)
    return item


@router.post("/projects/{project_id}/boq/upload-excel")
async def upload_boq_excel(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(_get_db),
):
    """Upload BOQ from Excel file and parse it"""
    import openpyxl
    from app.models.tenant.ipc_models import ProjectIPCSettings, BOQSection, BOQItem

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    if not settings:
        raise HTTPException(404, "Configure IPC settings first")

    # Clear existing BOQ
    for section in db.query(BOQSection).filter(BOQSection.settings_id == settings.id).all():
        db.delete(section)
    db.flush()

    current_section = None
    items_created = 0
    sections_created = 0
    total_value = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        row = list(row) + [None] * 10  # pad to avoid index errors

        # Detect section header (no quantity/rate, bold or empty unit)
        item_num = str(row[0] or "").strip()
        description = str(row[1] or "").strip()
        unit = str(row[2] or "").strip()
        qty_raw = row[3]
        rate_raw = row[4]

        if not description:
            continue

        # Section detection: no unit and no numeric quantity
        is_section = (not unit and qty_raw is None) or \
                     (item_num.isalpha() and len(item_num) == 1)

        if is_section:
            current_section = BOQSection(
                settings_id=settings.id,
                code=item_num or str(sections_created + 1),
                title=description,
                title_ar=description,
                sort_order=sections_created,
            )
            db.add(current_section)
            db.flush()
            sections_created += 1
        else:
            if not current_section:
                # Create default section
                current_section = BOQSection(
                    settings_id=settings.id,
                    code="A",
                    title="الأعمال العامة",
                    title_ar="الأعمال العامة",
                    sort_order=0,
                )
                db.add(current_section)
                db.flush()
                sections_created += 1

            try:
                qty = float(qty_raw) if qty_raw is not None else 0.0
                rate = float(rate_raw) if rate_raw is not None else 0.0
            except (TypeError, ValueError):
                qty, rate = 0.0, 0.0

            total = round(qty * rate, 2)
            total_value += total

            item = BOQItem(
                section_id=current_section.id,
                item_number=item_num,
                description=description,
                description_ar=description,
                unit=unit,
                quantity=qty,
                unit_rate=rate,
                total_amount=total,
                sort_order=items_created,
            )
            db.add(item)
            items_created += 1

    db.commit()
    return {
        "success": True,
        "sections_created": sections_created,
        "items_created": items_created,
        "total_contract_value": round(total_value, 2),
        "message": f"تم استيراد {items_created} بند في {sections_created} قسم",
    }


# ════════════════════════════════════════════════════════
# IPC CREATION & MANAGEMENT
# ════════════════════════════════════════════════════════

@router.get("/projects/{project_id}/ipcs")
def list_ipcs(project_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC
    return db.query(IPC).filter(
        IPC.project_id == project_id
    ).order_by(IPC.period_from.desc()).all()


@router.get("/ipcs/{ipc_id}")
def get_ipc(ipc_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404, "IPC not found")
    return {
        "ipc": ipc,
        "line_items": ipc.line_items,
        "summary": _ipc_summary(ipc),
    }


def _ipc_summary(ipc) -> dict:
    return {
        "gross_this_period": ipc.gross_this_period,
        "gross_cumulative": ipc.gross_cumulative,
        "retention_this": ipc.retention_this,
        "advance_recovery_this": ipc.advance_recovery_this,
        "previous_certificates": ipc.previous_certificates,
        "net_this_period": ipc.net_this_period,
        "vat_amount": ipc.vat_amount,
        "total_due": ipc.total_due,
    }


@router.post("/projects/{project_id}/ipcs/create")
def create_ipc(
    project_id: int,
    period_from: date = Body(...),
    period_to: date = Body(...),
    ipc_number: str = Body(None),
    db: Session = Depends(_get_db),
):
    """Create a new IPC from BOQ cumulative data"""
    from app.models.tenant.ipc_models import (
        IPC, IPCLineItem, BOQSection, BOQItem, ProjectIPCSettings
    )

    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == project_id
    ).first()
    if not settings:
        raise HTTPException(404, "Configure IPC settings first")

    # Auto-number
    if not ipc_number:
        count = db.query(IPC).filter(IPC.project_id == project_id).count()
        ipc_number = f"IPC-{count+1:03d}"

    # Get previous IPC for cumulative tracking
    prev_ipc = db.query(IPC).filter(
        IPC.project_id == project_id,
        IPC.period_to < period_from,
        IPC.status.in_(["approved", "paid"]),
    ).order_by(IPC.period_to.desc()).first()

    prev_cumulative = prev_ipc.gross_cumulative if prev_ipc else 0.0
    prev_retention = prev_ipc.retention_cumulative if prev_ipc else 0.0
    prev_advance = prev_ipc.advance_recovery_cumulative if prev_ipc else 0.0

    # Create IPC
    ipc = IPC(
        project_id=project_id,
        ipc_number=ipc_number,
        ipc_title=f"مستخلص {ipc_number} — {_month_name_ar(period_from.month)} {period_from.year}",
        ipc_title_ar=f"مستخلص رقم {ipc_number}",
        period_from=period_from,
        period_to=period_to,
        cutoff_date=period_to,
        status="draft",
        retention_pct=settings.retention_percentage,
        vat_pct=settings.vat_percentage,
        previous_certificates=prev_cumulative,
        is_auto_generated=True,
    )
    db.add(ipc)
    db.flush()

    # Create line items from BOQ
    total_this_period = 0.0
    sections = db.query(BOQSection).filter(
        BOQSection.settings_id == settings.id,
        BOQSection.is_active == True,
    ).all()

    sort_order = 0
    for section in sections:
        items = db.query(BOQItem).filter(
            BOQItem.section_id == section.id
        ).order_by(BOQItem.sort_order).all()

        for boq_item in items:
            prev_qty = boq_item.previous_qty or 0.0
            prev_amount = round(prev_qty * boq_item.unit_rate, 2)

            # Default this period: 0 (user will fill in)
            this_qty = 0.0
            this_amount = 0.0
            cum_qty = prev_qty + this_qty
            cum_amount = prev_amount + this_amount
            completion = round((cum_qty / boq_item.quantity * 100) if boq_item.quantity else 0, 1)

            line = IPCLineItem(
                ipc_id=ipc.id,
                boq_item_id=boq_item.id,
                item_number=boq_item.item_number,
                description=boq_item.description,
                description_ar=boq_item.description_ar,
                unit=boq_item.unit,
                contract_quantity=boq_item.quantity,
                previous_quantity=prev_qty,
                this_period_quantity=this_qty,
                cumulative_quantity=cum_qty,
                unit_rate=boq_item.unit_rate,
                previous_amount=prev_amount,
                this_period_amount=this_amount,
                cumulative_amount=cum_amount,
                completion_pct=completion,
                sort_order=sort_order,
            )
            db.add(line)
            sort_order += 1

    # Recalculate totals
    _recalculate_ipc_totals(ipc, prev_cumulative, prev_retention, prev_advance, settings)
    db.commit()
    db.refresh(ipc)
    return {"ipc_id": ipc.id, "ipc_number": ipc.ipc_number, "status": ipc.status}


def _recalculate_ipc_totals(ipc, prev_cum, prev_ret, prev_adv, settings):
    """Recalculate all IPC financial totals"""
    gross_this = sum(li.this_period_amount for li in ipc.line_items)
    gross_cum = prev_cum + gross_this
    ret_this = round(gross_this * settings.retention_percentage / 100, 2)
    ret_cum = prev_ret + ret_this
    # Advance recovery: recover proportionally
    if settings.contract_value > 0:
        adv_rate = settings.advance_amount / settings.contract_value
    else:
        adv_rate = settings.advance_payment_percentage / 100
    adv_this = round(gross_this * adv_rate, 2)
    adv_cum = prev_adv + adv_this
    net = round(gross_this - ret_this - adv_this, 2)
    vat = round(net * settings.vat_percentage / 100, 2)
    total = round(net + vat, 2)

    ipc.gross_this_period = gross_this
    ipc.gross_cumulative = gross_cum
    ipc.retention_this = ret_this
    ipc.retention_cumulative = ret_cum
    ipc.advance_recovery_this = adv_this
    ipc.advance_recovery_cumulative = adv_cum
    ipc.net_this_period = net
    ipc.vat_amount = vat
    ipc.total_due = total


@router.put("/ipcs/{ipc_id}/line-items/{line_id}")
def update_line_item(
    ipc_id: int,
    line_id: int,
    this_period_quantity: float = Body(...),
    notes: str = Body(""),
    db: Session = Depends(_get_db),
):
    """Update a line item's this-period quantity and recalculate"""
    from app.models.tenant.ipc_models import IPC, IPCLineItem, ProjectIPCSettings

    line = db.query(IPCLineItem).filter(
        IPCLineItem.id == line_id,
        IPCLineItem.ipc_id == ipc_id,
    ).first()
    if not line:
        raise HTTPException(404)

    line.this_period_quantity = this_period_quantity
    line.this_period_amount = round(this_period_quantity * line.unit_rate, 2)
    line.cumulative_quantity = line.previous_quantity + this_period_quantity
    line.cumulative_amount = line.previous_amount + line.this_period_amount
    if line.contract_quantity:
        line.completion_pct = round(line.cumulative_quantity / line.contract_quantity * 100, 1)
    line.notes = notes

    # Recalculate IPC totals
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    settings = db.query(ProjectIPCSettings).filter(
        ProjectIPCSettings.project_id == ipc.project_id
    ).first()
    prev_cum = ipc.previous_certificates
    _recalculate_ipc_totals(ipc, prev_cum, ipc.retention_cumulative - ipc.retention_this,
                             ipc.advance_recovery_cumulative - ipc.advance_recovery_this, settings)
    db.commit()
    return {"line_item": line, "ipc_totals": _ipc_summary(ipc)}


@router.post("/ipcs/{ipc_id}/submit")
def submit_ipc(ipc_id: int, notes: str = Body(""), db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC, IPCStatus
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)
    if ipc.status != IPCStatus.draft:
        raise HTTPException(400, f"Cannot submit IPC in status: {ipc.status}")
    ipc.status = IPCStatus.submitted
    ipc.submitted_at = datetime.utcnow()
    ipc.contractor_notes = notes
    db.commit()
    return {"status": "submitted", "ipc_number": ipc.ipc_number}


@router.post("/ipcs/{ipc_id}/approve")
def approve_ipc(
    ipc_id: int,
    approver_role: str = Body(...),  # "engineer" or "pm"
    approved_amount: float = Body(None),
    notes: str = Body(""),
    db: Session = Depends(_get_db),
):
    from app.models.tenant.ipc_models import IPC, IPCStatus
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)

    now = datetime.utcnow()
    if approver_role == "engineer":
        ipc.engineer_approved_at = now
        if ipc.status == IPCStatus.submitted:
            ipc.status = IPCStatus.under_review
        ipc.engineer_notes = notes
    elif approver_role == "pm":
        ipc.pm_approved_at = now
        ipc.status = IPCStatus.approved
        if approved_amount:
            ipc.total_due = approved_amount

    db.commit()
    return {"status": ipc.status, "ipc_number": ipc.ipc_number}


@router.post("/ipcs/{ipc_id}/mark-paid")
def mark_ipc_paid(
    ipc_id: int,
    amount_paid: float = Body(...),
    payment_date: date = Body(...),
    payment_reference: str = Body(""),
    db: Session = Depends(_get_db),
):
    from app.models.tenant.ipc_models import IPC, IPCStatus, BOQItem
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)

    ipc.amount_paid = amount_paid
    ipc.payment_date = payment_date
    ipc.payment_reference = payment_reference
    ipc.status = IPCStatus.paid if amount_paid >= ipc.total_due * 0.99 else IPCStatus.partially_paid

    # Update BOQ cumulative quantities
    for line in ipc.line_items:
        if line.boq_item_id:
            boq_item = db.query(BOQItem).filter(BOQItem.id == line.boq_item_id).first()
            if boq_item:
                boq_item.previous_qty = line.cumulative_quantity
                boq_item.previous_amount = line.cumulative_amount

    db.commit()
    return {"status": ipc.status, "amount_paid": amount_paid}


@router.post("/ipcs/{ipc_id}/reject")
def reject_ipc(ipc_id: int, reason: str = Body(...), db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC, IPCStatus
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)
    ipc.status = IPCStatus.rejected
    ipc.rejection_reason = reason
    db.commit()
    return {"status": "rejected"}


# ════════════════════════════════════════════════════════
# FILE UPLOAD (Excel/PDF)
# ════════════════════════════════════════════════════════

@router.post("/ipcs/{ipc_id}/upload")
async def upload_ipc_file(
    ipc_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(_get_db),
):
    """Upload IPC Excel or PDF and parse line items"""
    import openpyxl
    from app.models.tenant.ipc_models import IPC, IPCLineItem

    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)

    content = await file.read()
    filename = file.filename or ""
    parsed_items = []

    if filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or all(v is None for v in row):
                continue
            row = list(row) + [None] * 10
            desc = str(row[1] or "").strip()
            if not desc:
                continue
            try:
                this_qty = float(row[5] or 0)
                rate = float(row[3] or 0)
            except:
                continue

            parsed_items.append({
                "item_number": str(row[0] or ""),
                "description": desc,
                "unit": str(row[2] or ""),
                "this_period_quantity": this_qty,
                "unit_rate": rate,
                "this_period_amount": round(this_qty * rate, 2),
            })

        # Update matching line items
        updated = 0
        for parsed in parsed_items:
            for line in ipc.line_items:
                if line.item_number == parsed["item_number"] or \
                   (line.description and parsed["description"] in line.description):
                    line.this_period_quantity = parsed["this_period_quantity"]
                    line.this_period_amount = parsed["this_period_amount"]
                    line.cumulative_quantity = line.previous_quantity + line.this_period_quantity
                    line.cumulative_amount = line.previous_amount + line.this_period_amount
                    updated += 1
                    break

        ipc.uploaded_excel_path = f"uploads/ipc_{ipc_id}_{filename}"
        db.commit()
        return {
            "success": True,
            "file_type": "excel",
            "items_parsed": len(parsed_items),
            "items_updated": updated,
        }

    elif filename.endswith(".pdf"):
        ipc.uploaded_pdf_path = f"uploads/ipc_{ipc_id}_{filename}"
        db.commit()
        return {"success": True, "file_type": "pdf", "message": "PDF uploaded for reference"}

    raise HTTPException(400, "Unsupported file type. Use .xlsx or .pdf")


# ════════════════════════════════════════════════════════
# PDF EXPORT
# ════════════════════════════════════════════════════════

@router.get("/ipcs/{ipc_id}/export-pdf")
def export_ipc_pdf(ipc_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)

    pdf_bytes = _generate_ipc_pdf(ipc)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={ipc.ipc_number}.pdf"},
    )


@router.get("/ipcs/{ipc_id}/export-excel")
def export_ipc_excel(ipc_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPC
    ipc = db.query(IPC).filter(IPC.id == ipc_id).first()
    if not ipc:
        raise HTTPException(404)

    excel_bytes = _generate_ipc_excel(ipc)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={ipc.ipc_number}.xlsx"},
    )


def _generate_ipc_pdf(ipc) -> bytes:
    """Generate professional IPC PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfbase import pdfmetrics

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#0A1628")
    BLUE = colors.HexColor("#1565C0")
    LIGHT = colors.HexColor("#E3F2FD")
    GOLD = colors.HexColor("#FFB300")
    GREEN = colors.HexColor("#1B5E20")

    title_s = ParagraphStyle("tt", parent=styles["Title"], textColor=NAVY,
                              fontSize=14, alignment=TA_CENTER)
    sub_s = ParagraphStyle("sub", parent=styles["Normal"], textColor=BLUE,
                            fontSize=11, alignment=TA_CENTER)
    normal = styles["Normal"]
    bold = ParagraphStyle("bold", parent=normal, fontName="Helvetica-Bold")
    small = ParagraphStyle("small", parent=normal, fontSize=8)

    story = []

    # ── Header ──────────────────────────────────────────────
    story.append(Paragraph("INTERIM PAYMENT CERTIFICATE", title_s))
    story.append(Paragraph("شهادة الدفع المؤقتة — مستخلص", title_s))
    story.append(Spacer(1, 0.3*cm))

    # IPC Info table
    info = [
        [Paragraph("IPC Number / رقم المستخلص", bold), ipc.ipc_number,
         Paragraph("Status / الحالة", bold), ipc.status.upper()],
        [Paragraph("Period From / من", bold),
         str(ipc.period_from), Paragraph("Period To / إلى", bold), str(ipc.period_to)],
        [Paragraph("Project / المشروع", bold),
         ipc.ipc_title_ar or "", Paragraph("Cutoff Date / تاريخ القطع", bold),
         str(ipc.cutoff_date or "")],
    ]
    info_table = Table(info, colWidths=[5*cm, 8*cm, 5*cm, 8*cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT),
        ("BACKGROUND", (2, 0), (2, -1), LIGHT),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4*cm))

    # ── Line Items Table ──────────────────────────────────
    headers = [
        "Item\nرقم", "Description\nالوصف", "Unit\nالوحدة",
        "Contract Qty\nكمية العقد", "Prev. Qty\nكمية سابقة",
        "This Period\nهذه الفترة", "Cumulative\nتراكمي",
        "Rate\nالسعر", "Prev. Amt\nمبلغ سابق",
        "This Amt\nهذه الفترة", "Cum. Amt\nتراكمي",
        "Comp%\nنسبة"
    ]

    data = [headers]
    for line in sorted(ipc.line_items, key=lambda x: x.sort_order):
        data.append([
            line.item_number or "",
            Paragraph(line.description or "", small),
            line.unit or "",
            f"{line.contract_quantity:,.2f}" if line.contract_quantity else "—",
            f"{line.previous_quantity:,.2f}",
            f"{line.this_period_quantity:,.2f}",
            f"{line.cumulative_quantity:,.2f}",
            f"{line.unit_rate:,.2f}",
            f"{line.previous_amount:,.2f}",
            f"{line.this_period_amount:,.2f}",
            f"{line.cumulative_amount:,.2f}",
            f"{line.completion_pct:.1f}%",
        ])

    col_w = [1.2*cm, 6*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm,
             2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.5*cm]
    items_table = Table(data, colWidths=col_w, repeatRows=1)
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("PADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Financial Summary ─────────────────────────────────
    story.append(Paragraph("Financial Summary / الملخص المالي", sub_s))
    summary_data = [
        ["Description / البيان", "This Period / هذه الفترة", "Cumulative / تراكمي"],
        ["Gross Work Done / قيمة الأعمال",
         f"SAR {ipc.gross_this_period:,.2f}", f"SAR {ipc.gross_cumulative:,.2f}"],
        [f"Retention ({ipc.retention_pct}%) / الضمان",
         f"(SAR {ipc.retention_this:,.2f})", f"(SAR {ipc.retention_cumulative:,.2f})"],
        ["Advance Recovery / استرداد مقدم",
         f"(SAR {ipc.advance_recovery_this:,.2f})",
         f"(SAR {ipc.advance_recovery_cumulative:,.2f})"],
        ["Previous Certificates / شهادات سابقة", "—",
         f"(SAR {ipc.previous_certificates:,.2f})"],
        ["Net This Period / صافي الفترة",
         f"SAR {ipc.net_this_period:,.2f}", ""],
        [f"VAT ({ipc.vat_pct}%) / ضريبة القيمة المضافة",
         f"SAR {ipc.vat_amount:,.2f}", ""],
        ["TOTAL DUE / الإجمالي المستحق",
         f"SAR {ipc.total_due:,.2f}", ""],
    ]
    sum_table = Table(summary_data, colWidths=[8*cm, 6*cm, 6*cm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), GOLD),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(sum_table)

    # Signatures
    story.append(Spacer(1, 0.6*cm))
    sig_data = [
        [Paragraph("Contractor / المقاول", bold),
         Paragraph("Engineer / المهندس", bold),
         Paragraph("Project Manager / مدير المشروع", bold),
         Paragraph("Client / العميل", bold)],
        ["", "", "", ""],
    ]
    sig_table = Table(sig_data, colWidths=[7*cm, 7*cm, 7*cm, 7*cm])
    sig_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
        ("ROWHEIGHTS", (0, 1), (-1, 1), 40),
        ("PADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]))
    story.append(Paragraph("Signatures / التوقيعات", sub_s))
    story.append(sig_table)

    # Footer
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        f"OMMS v3.0 | Generated: {date.today()} | {ipc.ipc_number} | Confidential",
        ParagraphStyle("footer", parent=normal, fontSize=7.5,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(story)
    return buf.getvalue()


def _generate_ipc_excel(ipc) -> bytes:
    """Generate professional IPC Excel workbook"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    NAVY = "0A1628"
    BLUE = "1565C0"
    LIGHT = "E3F2FD"
    GOLD = "FFB300"
    RED = "C62828"
    GREEN = "1B5E20"
    WHITE = "FFFFFF"

    def hd(cell, bg=BLUE, fg=WHITE, bold=True, size=9, align="center"):
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    def border(cell):
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fmt(cell, val, fmt_str="#,##0.00"):
        cell.value = val
        cell.number_format = fmt_str
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: IPC Cover ────────────────────────────────
    ws = wb.active
    ws.title = "IPC Cover"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:L1")
    ws["A1"] = f"INTERIM PAYMENT CERTIFICATE — {ipc.ipc_number}"
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"] = f"شهادة الدفع المؤقتة | الفترة: {ipc.period_from} إلى {ipc.period_to}"
    ws["A2"].font = Font(bold=True, size=12, color=BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Info block
    info = [
        ("رقم المستخلص / IPC Number", ipc.ipc_number),
        ("الحالة / Status", ipc.status),
        ("من / Period From", str(ipc.period_from)),
        ("إلى / Period To", str(ipc.period_to)),
        ("تاريخ القطع / Cutoff", str(ipc.cutoff_date or "")),
    ]
    for i, (label, val) in enumerate(info, 4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = PatternFill(fill_type="solid", fgColor=LIGHT)
        ws.cell(row=i, column=2, value=val)
        for col in range(1, 3):
            border(ws.cell(row=i, column=col))

    # Line Items
    headers_row = 10
    hdrs = ["رقم", "الوصف", "الوحدة", "كمية العقد",
            "كمية سابقة", "كمية الفترة", "تراكمي",
            "سعر الوحدة", "مبلغ سابق", "مبلغ الفترة", "تراكمي", "نسبة الإنجاز"]
    col_widths = [8, 35, 8, 14, 14, 14, 14, 14, 15, 15, 15, 12]

    for col, (h, w) in enumerate(zip(hdrs, col_widths), 1):
        cell = ws.cell(row=headers_row, column=col, value=h)
        hd(cell, bg=NAVY)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[headers_row].height = 30

    for r, line in enumerate(sorted(ipc.line_items, key=lambda x: x.sort_order), headers_row + 1):
        row_data = [
            line.item_number, line.description, line.unit,
            line.contract_quantity, line.previous_quantity,
            line.this_period_quantity, line.cumulative_quantity,
            line.unit_rate, line.previous_amount,
            line.this_period_amount, line.cumulative_amount,
            line.completion_pct / 100,
        ]
        fill_color = LIGHT if r % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            border(cell)
            if col in [4, 5, 6, 7]:
                cell.number_format = "#,##0.000"
            elif col in [8, 9, 10, 11]:
                cell.number_format = "#,##0.00"
            elif col == 12:
                cell.number_format = "0.0%"

    # Summary
    last_item_row = headers_row + len(ipc.line_items) + 2
    summary = [
        ("قيمة الأعمال هذه الفترة", ipc.gross_this_period, LIGHT),
        (f"خصم الضمان ({ipc.retention_pct}%)", -ipc.retention_this, "FFEBEE"),
        ("استرداد الدفعة المقدمة", -ipc.advance_recovery_this, "FFEBEE"),
        ("شهادات سابقة", -ipc.previous_certificates, "FFEBEE"),
        ("صافي هذه الفترة", ipc.net_this_period, "E8F5E9"),
        (f"ضريبة القيمة المضافة ({ipc.vat_pct}%)", ipc.vat_amount, LIGHT),
        ("الإجمالي المستحق", ipc.total_due, GOLD),
    ]
    for i, (label, val, color) in enumerate(summary):
        r = last_item_row + i
        label_cell = ws.cell(row=r, column=9, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(fill_type="solid", fgColor=color)
        val_cell = ws.cell(row=r, column=11, value=val)
        val_cell.number_format = "#,##0.00"
        val_cell.font = Font(bold=True, size=10)
        val_cell.fill = PatternFill(fill_type="solid", fgColor=color)
        for col in [9, 10, 11]:
            border(ws.cell(row=r, column=col))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════
# IPC NOTIFICATIONS SCHEDULER
# ════════════════════════════════════════════════════════

@router.post("/notifications/check-due")
def check_ipc_notifications(background_tasks: BackgroundTasks, db: Session = Depends(_get_db)):
    """Trigger IPC notification check (called by scheduler)"""
    background_tasks.add_task(_send_ipc_notifications, db)
    return {"message": "Notification check scheduled"}


def _send_ipc_notifications(db: Session):
    """Check for upcoming IPC deadlines and send notifications"""
    from app.models.tenant.ipc_models import ProjectIPCSettings, IPC, IPCNotification

    today = date.today()
    all_settings = db.query(ProjectIPCSettings).all()

    for settings in all_settings:
        schedule = _generate_ipc_schedule(settings, settings.project, today.year)

        for slot in schedule:
            due_date = date.fromisoformat(slot["due_date"])
            days_remaining = (due_date - today).days

            # Check if notification needed
            if days_remaining == settings.notify_days_before or \
               days_remaining in [3, 1, 0] or \
               days_remaining < 0:

                # Check if notification already sent today
                existing = db.query(IPCNotification).filter(
                    IPCNotification.project_id == settings.project_id,
                    IPCNotification.due_date == due_date,
                    IPCNotification.created_at >= datetime.utcnow().replace(
                        hour=0, minute=0, second=0),
                ).first()

                if existing:
                    continue

                if days_remaining < 0:
                    title = f"⚠️ مستخلص متأخر — {slot['ipc_number']}"
                    title_ar = f"المستخلص {slot['ipc_number']} تجاوز موعد التقديم"
                    ntype = "overdue"
                elif days_remaining == 0:
                    title = f"🔴 موعد رفع المستخلص اليوم — {slot['ipc_number']}"
                    title_ar = title
                    ntype = "due_today"
                else:
                    title = f"🔔 موعد المستخلص {slot['ipc_number']} خلال {days_remaining} يوم"
                    title_ar = title
                    ntype = "due_soon"

                notif = IPCNotification(
                    project_id=settings.project_id,
                    notification_type=ntype,
                    title=title,
                    title_ar=title_ar,
                    message=f"يستحق المستخلص {slot['ipc_number']} بتاريخ {due_date}",
                    message_ar=f"الفترة: {slot['period_from']} إلى {slot['period_to']}",
                    due_date=due_date,
                    sent_to=settings.notify_user_ids or [],
                )
                db.add(notif)

    db.commit()


@router.get("/projects/{project_id}/notifications")
def get_ipc_notifications(project_id: int, db: Session = Depends(_get_db)):
    from app.models.tenant.ipc_models import IPCNotification
    return db.query(IPCNotification).filter(
        IPCNotification.project_id == project_id,
    ).order_by(IPCNotification.created_at.desc()).limit(50).all()
