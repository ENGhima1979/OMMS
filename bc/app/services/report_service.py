"""
OMMS Report Generation Service
Generates PDF and Excel reports for work orders, KPIs, inventory, etc.
"""

from datetime import date, timedelta
from typing import Optional
import io


def generate_work_order_pdf(work_order_data: dict) -> bytes:
    """Generate a PDF work order card"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    BLUE = colors.HexColor("#185FA5")
    LIGHT_BLUE = colors.HexColor("#E6F1FB")
    ORANGE = colors.HexColor("#EF9F27")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=BLUE, fontSize=16, spaceAfter=4)
    header_style = ParagraphStyle("header", parent=styles["Normal"],
                                   textColor=colors.white, fontSize=10, fontName="Helvetica-Bold")
    normal = styles["Normal"]
    bold = ParagraphStyle("bold", parent=normal, fontName="Helvetica-Bold")

    story = []

    # Header
    story.append(Paragraph("OMMS — Work Order", title_style))
    story.append(Paragraph("نظام إدارة التشغيل والصيانة — أمر العمل", title_style))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
    story.append(Spacer(1, 0.3*cm))

    # WO Info table
    wo_info = [
        [Paragraph("WO Number / رقم أمر العمل", bold),
         Paragraph(str(work_order_data.get("wo_number", "")), normal),
         Paragraph("Status / الحالة", bold),
         Paragraph(str(work_order_data.get("status", "")), normal)],
        [Paragraph("Type / النوع", bold),
         Paragraph(str(work_order_data.get("wo_type", "")), normal),
         Paragraph("Priority / الأولوية", bold),
         Paragraph(str(work_order_data.get("priority", "")), normal)],
        [Paragraph("Asset / الأصل", bold),
         Paragraph(str(work_order_data.get("asset_name", "")), normal),
         Paragraph("Scheduled / مجدول", bold),
         Paragraph(str(work_order_data.get("scheduled_date", "")), normal)],
        [Paragraph("Title / العنوان", bold),
         Paragraph(str(work_order_data.get("title", "")), normal),
         Paragraph("Est. Duration / المدة المقدرة", bold),
         Paragraph(f"{work_order_data.get('estimated_duration_hours', 0)} hrs", normal)],
    ]

    t = Table(wo_info, colWidths=[4*cm, 7*cm, 4*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
        ("BACKGROUND", (2, 0), (2, -1), LIGHT_BLUE),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Description
    if work_order_data.get("description"):
        story.append(Paragraph("Description / الوصف:", bold))
        story.append(Paragraph(work_order_data["description"], normal))
        story.append(Spacer(1, 0.3*cm))

    # Checklist
    checklist = work_order_data.get("checklist_items", [])
    if checklist:
        story.append(Paragraph("Checklist / قائمة المراجعة:", bold))
        cl_data = [["#", "Task / المهمة", "Status / الحالة", "Notes / ملاحظات"]]
        for i, item in enumerate(checklist, 1):
            task = item.get("item", "") if isinstance(item, dict) else str(item)
            task_ar = item.get("item_ar", "") if isinstance(item, dict) else ""
            cl_data.append([
                str(i),
                f"{task}\n{task_ar}",
                "☐ Done",
                "",
            ])
        cl_table = Table(cl_data, colWidths=[1*cm, 9*cm, 4*cm, 5*cm])
        cl_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(cl_table)
        story.append(Spacer(1, 0.5*cm))

    # Signatures
    sig_data = [
        [Paragraph("Technician Signature\nتوقيع الفني", bold),
         Paragraph("Supervisor Signature\nتوقيع المشرف", bold),
         Paragraph("Completion Date\nتاريخ الإنجاز", bold)],
        ["", "", ""],
    ]
    sig_table = Table(sig_data, colWidths=[6*cm, 6*cm, 7*cm])
    sig_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT_BLUE),
        ("ROWHEIGHTS", (0, 1), (-1, 1), 50),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(Paragraph("Completion & Signatures / الإنجاز والتوقيعات:", bold))
    story.append(sig_table)

    # Footer
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        f"Generated by OMMS v2.0 | {date.today()} | Confidential",
        ParagraphStyle("footer", parent=normal, fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_maintenance_report_excel(data: dict) -> bytes:
    """Generate Excel maintenance summary report"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    BLUE = "185FA5"
    LIGHT_BLUE = "E6F1FB"
    ORANGE = "EF9F27"
    GREEN = "1D9E75"

    def style_header(cell, bg=BLUE):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def border_cell(cell):
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Summary ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary | ملخص"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = "OMMS — Maintenance Performance Report / تقرير أداء الصيانة"
    ws1["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1["A2"] = f"Period: {data.get('period_from', '')} to {data.get('period_to', '')}"
    ws1["A2"].font = Font(italic=True, color="666666")
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[3].height = 25

    # KPI table
    kpi_headers = ["KPI / المؤشر", "Target / المستهدف", "Actual / الفعلي", "Status / الحالة"]
    for col, h in enumerate(kpi_headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = 22

    kpis = [
        ("WO Completion Rate / معدل إنجاز أوامر العمل", "95%",
         f"{data.get('completion_rate', 0):.1f}%",
         "✅" if data.get('completion_rate', 0) >= 95 else "⚠️"),
        ("Asset Availability / توافر الأصول", "98%",
         f"{data.get('asset_availability', 0):.1f}%",
         "✅" if data.get('asset_availability', 0) >= 98 else "⚠️"),
        ("MTTR / متوسط وقت الإصلاح", "< 4 hrs",
         f"{data.get('mttr', 0):.1f} hrs",
         "✅" if data.get('mttr', 0) <= 4 else "⚠️"),
        ("OEE / الكفاءة الشاملة للمعدات", "> 85%",
         f"{data.get('oee', 0):.1f}%",
         "✅" if data.get('oee', 0) >= 85 else "⚠️"),
        ("Budget Utilization / استخدام الميزانية", "< 100%",
         f"{data.get('budget_used_pct', 0):.1f}%",
         "✅" if data.get('budget_used_pct', 0) <= 100 else "🚨"),
        ("Open Incidents / الحوادث المفتوحة", "0",
         str(data.get('open_incidents', 0)),
         "✅" if data.get('open_incidents', 0) == 0 else "🚨"),
    ]

    for row_idx, (kpi, target, actual, status) in enumerate(kpis, 4):
        row_data = [kpi, target, actual, status]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
            border_cell(cell)

    # ── Sheet 2: Work Orders ──────────────────────────────
    ws2 = wb.create_sheet("Work Orders | أوامر العمل")
    wo_headers = ["WO# | رقم أمر العمل", "Title | العنوان", "Type | النوع",
                  "Status | الحالة", "Priority | الأولوية", "Scheduled | مجدول",
                  "Duration (hrs) | المدة", "Cost (SAR) | التكلفة"]
    for col, h in enumerate(wo_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        style_header(cell)
        ws2.column_dimensions[get_column_letter(col)].width = 20

    for row_idx, wo in enumerate(data.get("work_orders", []), 2):
        row_vals = [
            wo.get("wo_number"), wo.get("title"), wo.get("wo_type"),
            wo.get("status"), wo.get("priority"), str(wo.get("scheduled_date", "")),
            wo.get("actual_duration_hours"), wo.get("actual_cost"),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
            border_cell(cell)

    # ── Sheet 3: Inventory ────────────────────────────────
    ws3 = wb.create_sheet("Inventory | المخزون")
    inv_headers = ["Part# | رقم القطعة", "Name | الاسم", "Stock | المخزون",
                   "Min Qty | الحد الأدنى", "Reorder | نقطة الطلب",
                   "Unit Cost | تكلفة الوحدة", "Total Value | القيمة الكلية", "Status | الحالة"]
    for col, h in enumerate(inv_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        style_header(cell, bg=GREEN)
        ws3.column_dimensions[get_column_letter(col)].width = 20

    for row_idx, part in enumerate(data.get("spare_parts", []), 2):
        qty = part.get("quantity_in_stock", 0)
        reorder = part.get("reorder_point", 0)
        status = "🔴 Low" if qty <= reorder else "🟢 OK"
        row_vals = [
            part.get("part_number"), part.get("name"), qty,
            part.get("minimum_quantity"), reorder,
            part.get("unit_cost"), part.get("total_value"), status,
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if qty <= reorder:
                cell.fill = PatternFill(fill_type="solid", fgColor="FDECEA")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
            border_cell(cell)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
